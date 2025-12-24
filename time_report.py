import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from datetime import datetime, time, timedelta
import os


def setup_logging():
    """Configure logging to file and console"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('time_report.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger()


def calculate_hours(start_str, end_str, logger):
    """Calculate regular (6am-6pm) and night (6pm-6am) hours between two timestamps."""
    try:
        start_dt = pd.to_datetime(start_str, format='%m/%d/%Y %I:%M:%S%p')
        end_dt = pd.to_datetime(end_str, format='%m/%d/%Y %I:%M:%S%p')
        
        if start_dt > end_dt:
            logger.warning(f"Start time {start_str} is after end time {end_str}")
            return 0.0, 0.0
            
        regular_hours = 0.0
        night_hours = 0.0
        current = start_dt
        
        while current < end_dt:
            day_start = datetime.combine(current.date(), time(6, 0))   # 6:00 AM
            day_end = datetime.combine(current.date(), time(18, 0))    # 6:00 PM
            
            if current < day_start:
                # Current time is before 6am (night)
                segment_end = min(day_start, end_dt)
                night_hours += (segment_end - current).total_seconds() / 3600
                current = segment_end
            elif current < day_end:
                # Current time is between 6am and 6pm (regular)
                segment_end = min(day_end, end_dt)
                regular_hours += (segment_end - current).total_seconds() / 3600
                current = segment_end
            else:
                # Current time is after 6pm (night)
                next_day_start = day_start + timedelta(days=1)
                segment_end = min(next_day_start, end_dt)
                night_hours += (segment_end - current).total_seconds() / 3600
                current = segment_end
                
        return round(regular_hours, 5), round(night_hours, 5)
    except Exception as e:
        logger.error(f"Error calculating hours for {start_str} to {end_str}: {e}")
        return 0.0, 0.0


def apply_table_formatting(worksheet, start_row, start_col, end_row, end_col, table_name, header_color):
    """Apply formatting to Excel table."""
    # Header formatting
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def apply_number_formatting(worksheet, start_row, end_row, start_col, end_col):
    """Apply number formatting to specified columns."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00000'
                cell.alignment = Alignment(horizontal="right")


def add_total_row_formatting(worksheet, row, col_count, fill_color="D3D3D3"):
    """Apply special formatting to total row."""
    total_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    total_font = Font(bold=True)
    
    for col in range(1, col_count + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = total_font


def autofit_columns(worksheet):
    """Auto-fit columns in Excel worksheet."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def process_time_report(input_file, output_file):
    """Process the time tracking data and generate reports."""
    logger = setup_logging()
    
    try:
        # Read input file
        logger.info(f"Reading input file: {input_file}")
        df = pd.read_excel(input_file) if input_file.endswith('.xlsx') else pd.read_csv(input_file)
        
        # Log total agents
        total_agents = df['Employee Name'].nunique()
        logger.info(f"Total agents in dataset: {total_agents}")
        
        # Define categories of interest
        categories = [
            'Logged In',
            'Logged Out',
            'Meeting',
            'Meeting - Pre-Shift',
            'Overtime Withdrawals',
            'Withdrawals',
            'Wrap-Up'
        ]
        
        # Filter for selected categories and exclude Training queue
        filtered_df = df[
            df['Actual State'].isin(categories) & 
            (df['Queue'] != 'Training')
        ].copy()
        logger.info(f"Filtered to {len(filtered_df)} records in specified categories (excluding Training queue)")
        
        # Extract date from Start time with explicit format
        filtered_df['Date'] = pd.to_datetime(filtered_df['Start'], format='%m/%d/%Y %I:%M:%S%p').dt.date
        
        # Classify IBC vs Non-IBC
        def classify_queue(q):
            if pd.isna(q) or q == '':
                return 'Non IBC'
            return 'IBC' if 'ibc' in str(q).lower() else 'Non IBC'
            
        filtered_df['Queue_Type'] = filtered_df['Queue'].apply(classify_queue)
        
        # Calculate regular and night hours
        logger.info("Calculating regular and night hours")
        regular_hours_list = []
        night_hours_list = []
        
        for index, row in filtered_df.iterrows():
            regular_hrs, night_hrs = calculate_hours(row['Start'], row['Stop'], logger)
            regular_hours_list.append(regular_hrs)
            night_hours_list.append(night_hrs)
            
        filtered_df['Regular Hours'] = regular_hours_list
        filtered_df['Night Hours'] = night_hours_list
        filtered_df['Total Hours'] = filtered_df['Regular Hours'] + filtered_df['Night Hours']
        
        # Create Agent Summary
        logger.info("Creating Agent Summary")
        agent_summary = filtered_df.groupby([
            'Employee Name', 
            'Queue',
            'Date',
            'Actual State',
            'Queue_Type'
        ]).agg({
            'Regular Hours': 'sum',
            'Night Hours': 'sum',
            'Total Hours': 'sum'
        }).reset_index()
        
        # Separate IBC and Non-IBC summaries
        ibc_agent_summary = agent_summary[agent_summary['Queue_Type'] == 'IBC'].drop(columns='Queue_Type')
        non_ibc_agent_summary = agent_summary[agent_summary['Queue_Type'] == 'Non IBC'].drop(columns='Queue_Type')
        
        # Add totals
        ibc_total = pd.DataFrame({
            'Employee Name': ['TOTAL'],
            'Queue': [''],
            'Date': [''],
            'Actual State': [''],
            'Regular Hours': [ibc_agent_summary['Regular Hours'].sum()],
            'Night Hours': [ibc_agent_summary['Night Hours'].sum()],
            'Total Hours': [ibc_agent_summary['Total Hours'].sum()]
        })
        ibc_agent_summary = pd.concat([ibc_agent_summary, ibc_total], ignore_index=True)
        
        non_ibc_total = pd.DataFrame({
            'Employee Name': ['TOTAL'],
            'Queue': [''],
            'Date': [''],
            'Actual State': [''],
            'Regular Hours': [non_ibc_agent_summary['Regular Hours'].sum()],
            'Night Hours': [non_ibc_agent_summary['Night Hours'].sum()],
            'Total Hours': [non_ibc_agent_summary['Total Hours'].sum()]
        })
        non_ibc_agent_summary = pd.concat([non_ibc_agent_summary, non_ibc_total], ignore_index=True)
        
        # Create Category Summary
        logger.info("Creating Category Summary")
        category_summary = filtered_df.groupby([
            'Queue_Type',
            'Actual State'
        ]).agg({
            'Regular Hours': 'sum',
            'Night Hours': 'sum'
        }).reset_index()
        
        # Pivot to get IBC and Non-IBC side by side
        ibc_cat = category_summary[category_summary['Queue_Type'] == 'IBC'].drop(columns='Queue_Type')
        non_ibc_cat = category_summary[category_summary['Queue_Type'] == 'Non IBC'].drop(columns='Queue_Type')
        
        final_category = pd.merge(
            ibc_cat,
            non_ibc_cat,
            on='Actual State',
            how='outer',
            suffixes=('_IBC', '_Non_IBC')
        ).fillna(0)
        
        # Add totals row
        cat_total = pd.DataFrame({
            'Actual State': ['TOTAL'],
            'Regular Hours_IBC': [final_category['Regular Hours_IBC'].sum()],
            'Night Hours_IBC': [final_category['Night Hours_IBC'].sum()],
            'Regular Hours_Non_IBC': [final_category['Regular Hours_Non_IBC'].sum()],
            'Night Hours_Non_IBC': [final_category['Night Hours_Non_IBC'].sum()]
        })
        final_category = pd.concat([final_category, cat_total], ignore_index=True)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # IBC Agent Summary Sheet
        ws_ibc = wb.active
        ws_ibc.title = "IBC Agent Summary"
        ws_ibc.cell(row=1, column=1, value="IBC Agent Summary").font = Font(size=14, bold=True)
        
        # Write IBC agent data
        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(ibc_agent_summary, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws_ibc.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        end_row = start_row + len(ibc_agent_summary)
        apply_table_formatting(ws_ibc, start_row, 1, end_row - 1, 7, "IBC_Table", "4472C4")
        apply_number_formatting(ws_ibc, start_row + 1, end_row, 5, 7)  # Format number columns
        add_total_row_formatting(ws_ibc, end_row, 7)
        
        # Non-IBC Agent Summary Sheet
        ws_non_ibc = wb.create_sheet("Non-IBC Agent Summary")
        ws_non_ibc.cell(row=1, column=1, value="Non-IBC Agent Summary").font = Font(size=14, bold=True)
        
        # Write Non-IBC agent data
        start_row_non_ibc = 3
        for r_idx, row in enumerate(dataframe_to_rows(non_ibc_agent_summary, index=False, header=True), start_row_non_ibc):
            for c_idx, value in enumerate(row, 1):
                ws_non_ibc.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        end_row_non_ibc = start_row_non_ibc + len(non_ibc_agent_summary)
        apply_table_formatting(ws_non_ibc, start_row_non_ibc, 1, end_row_non_ibc - 1, 7, "NonIBC_Table", "ED7D31")
        apply_number_formatting(ws_non_ibc, start_row_non_ibc + 1, end_row_non_ibc, 5, 7)  # Format number columns
        add_total_row_formatting(ws_non_ibc, end_row_non_ibc, 7)
        
        # Category Summary Sheet
        ws_cat = wb.create_sheet("Category Summary")
        ws_cat.cell(row=1, column=1, value="Category Summary").font = Font(size=14, bold=True)
        
        # Write category data
        start_row_cat = 3
        for r_idx, row in enumerate(dataframe_to_rows(final_category, index=False, header=True), start_row_cat):
            for c_idx, value in enumerate(row, 1):
                ws_cat.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        cat_end_row = start_row_cat + len(final_category)
        apply_table_formatting(ws_cat, start_row_cat, 1, cat_end_row - 1, 5, "Category_Table", "70AD47")
        apply_number_formatting(ws_cat, start_row_cat + 1, cat_end_row, 2, 5)  # Format number columns
        add_total_row_formatting(ws_cat, cat_end_row, 5, "A9D08E")
        
        # Auto-fit columns
        for sheet in wb.sheetnames:
            autofit_columns(wb[sheet])
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Report saved to {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error processing report: {e}")
        raise


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python time_report.py <input_file> <output_file>")
        sys.exit(1)
        
    process_time_report(sys.argv[1], sys.argv[2])
