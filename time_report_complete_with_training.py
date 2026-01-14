import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from datetime import datetime, time, timedelta
import os
from fuzzywuzzy import fuzz, process
import re
import argparse

def setup_logging():
    """Configure logging to file and console"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('time_report_complete.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger()

# Functions from trepor_etl.py
def normalize_name(name):
    """Enhanced name normalization with fuzzy matching preparation"""
    if pd.isna(name) or not isinstance(name, str):
        return None, None
    
    original = name.strip()
    if not original:
        return None, None
    
    # Remove extra spaces and special characters
    cleaned = re.sub(r'[^a-zA-Z, ]', '', original).strip()
    
    # Handle 'Lastname, Firstname' format
    if ',' in cleaned:
        parts = [p.strip() for p in cleaned.split(',')]
        if len(parts) >= 2:
            # Return both formats for matching
            standard = f"{parts[1]} {parts[0]}".strip()
            return standard, original
        return cleaned, original
    
    # Handle 'Firstname Lastname' format
    return cleaned, original

def match_names(tracking_names, affil_names, min_confidence=80):
    """Fuzzy match names with a minimum confidence threshold"""
    matches = {}
    unmatched = []
    
    for t_name in tracking_names:
        if pd.isna(t_name):
            continue
            
        # Try exact match first
        if t_name in affil_names:
            matches[t_name] = t_name
            continue
            
        # Try fuzzy match
        best_match, score = process.extractOne(t_name, affil_names, scorer=fuzz.token_sort_ratio)
        if score >= min_confidence:
            matches[t_name] = best_match
        else:
            unmatched.append(t_name)
            logging.warning(f"No good match found for '{t_name}' (best: '{best_match}' at {score}%)")
            
    return matches, unmatched

def process_employee_data(input_file):
    """Process raw employee data without saving to file"""
    logger = setup_logging()
    logger.info(f"Starting processing of file: {input_file}")
    
    try:
        # Read the input file
        logger.info("Reading input file...")
        if input_file.lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(input_file, header=None, dtype=str)
        else:
            df = pd.read_csv(input_file, header=None, dtype=str, skip_blank_lines=False)
       
        df=df.dropna(how='all')
        df=df.dropna(axis='columns', how='all')
        logger.info(f"Successfully read file with {len(df)} rows")
    except Exception as e:
        logger.error(f"Error reading input file: {e}")
        return

    # Initialize variables
    employee_group = None
    employee_name = None
    records = []
    
    # Process each row
    for index, row in df.iterrows():
        # Convert row to list and clean up
        row_data = [str(x).strip() if pd.notna(x) and str(x).strip() != 'nan' else '' for x in row]
        
        # Remove empty columns from the row
        non_empty_cols = [i for i, val in enumerate(row_data) if val]
        if non_empty_cols:
            min_col, max_col = min(non_empty_cols), max(non_empty_cols)
            row_data = row_data[min_col:max_col+1]
        else:
            row_data = []
            
        # Skip empty rows
        if not any(row_data):
            continue
            
        # Check for Employee Group
        if any('Employee Group:' in x for x in row_data):
            employee_group = ' '.join(row_data).split('Employee Group:')[-1].strip().split(':')[0].strip()
            logger.info(f"Found Employee Group: {employee_group}")
            continue
            
        # Check for Employee Name
        if any('Employee Name:' in x for x in row_data):
            employee_name = ' '.join(row_data).split('Employee Name:')[-1].strip()
            logger.info(f"Found Employee Name: {employee_name}")
            continue
            
        # Check if this row matches the data pattern (date values present)
        if any('/' in x for x in row_data) and len(row_data) >= 7:
            try:
                # Extract data fields
                record = {
                    'Employee Group': employee_group,
                    'Employee Name': employee_name,
                    'Start': row_data[0] if len(row_data) > 0 else '',
                    'Stop': row_data[2] if len(row_data) > 1 else '',
                    'Duration': row_data[3] if len(row_data) > 2 else '',
                    'In Schedule': row_data[4] if len(row_data) > 3 else '',
                    'In Adherence': row_data[5] if len(row_data) > 4 else '',
                    'Scheduled State': row_data[6] if len(row_data) > 5 else '',
                    'Actual State': row_data[7] if len(row_data) > 6 else ''
                }
                records.append(record)
                logger.debug(f"Added record: {record}")
                
            except Exception as e:
                logger.warning(f"Row {index+1}: Error processing data - {e}")
                continue
    
    # Create output DataFrame
    if records:
        output_df = pd.DataFrame(records)
        
        # Reorder columns to have Employee Group and Name first
        column_order = ['Employee Group', 'Employee Name'] + [col for col in output_df.columns if col not in ['Employee Group', 'Employee Name']]
        output_df = output_df[column_order]
        
        logger.info(f"Created output with {len(output_df)} records")
        return output_df
    else:
        logger.warning("No valid records found in the input file")
        return None

def add_queue_to_tracking_df(tracking_df, affiliation_file, min_confidence=80):
    """Add queue information to tracking data without saving intermediate files"""
    logger = setup_logging()
    logger.info("Starting enhanced queue processing with fuzzy matching")
    
    try:
        logger.info(f"Reading affiliation file: {affiliation_file}")
        # Read the roster Excel file, specifically the "7MS Main Roster" sheet
        affiliation_df = pd.read_excel(affiliation_file, sheet_name='7MS Main Roster ')
        
        # Normalize names in both dataframes
        logger.info("Normalizing employee names")
        tracking_df[['Normalized_Name', 'Original_Tracking_Name']] = tracking_df['Employee Name'].apply(
            lambda x: pd.Series(normalize_name(x))
        )
        affiliation_df[['Normalized_Name', 'Original_Affil_Name']] = affiliation_df['Name'].apply(
            lambda x: pd.Series(normalize_name(x))
        )
        
        # Get unique normalized names
        tracking_names = tracking_df['Normalized_Name'].dropna().unique()
        affil_names = affiliation_df['Normalized_Name'].dropna().unique()
        
        # Match names with fuzzy logic
        logger.info("Matching names with fuzzy logic")
        name_matches, unmatched = match_names(tracking_names, affil_names, min_confidence)
        
        # Report matching results
        logger.info(f"Matched {len(name_matches)} out of {len(tracking_names)} names")
        if unmatched:
            logger.warning(f"Could not match {len(unmatched)} names: {unmatched}")
        
        # Create mapping dictionary
        name_map = {k: v for k, v in name_matches.items()}
        
        # Map the matched names
        tracking_df['Matched_Name'] = tracking_df['Normalized_Name'].map(name_map)
        
        # Merge dataframes
        logger.info("Merging data")
        merged_df = pd.merge(
            tracking_df,
            affiliation_df[['Normalized_Name', 'Supervisor', 'Manager', 'Department', 'Role', 'Shift', 'Schedule', 'Batch']],
            left_on='Matched_Name',
            right_on='Normalized_Name',
            how='left'
        )
        
        # Clean up columns
        merged_df = merged_df.drop(columns=['Normalized_Name_x', 'Normalized_Name_y', 'Matched_Name'])
        
        # Reorder columns to include additional fields
        columns = ['Employee Name', 'Batch', 'Supervisor', 'Manager', 'Department', 'Role', 'Shift', 'Schedule', 
                  'Start', 'Stop', 'Duration', 'In Schedule', 'In Adherence', 'Scheduled State', 'Actual State']
        # Keep only existing columns
        columns = [col for col in columns if col in merged_df.columns]
        merged_df = merged_df[columns]
        merged_df = merged_df.rename(columns={'Department': 'Queue'})

        # Generate matching report
        match_rate = len(name_matches) / len(tracking_names) * 100
        logger.info(f"Name matching completed with {match_rate:.2f}% success rate")
        
        return merged_df, match_rate
        
    except Exception as e:
        logger.error(f"Error during processing: {str(e)}", exc_info=True)
        raise

# Functions from time_report.py
def calculate_hours(start_str, end_str, logger):
    """Calculate regular (6am-6pm) and night (6pm-6am) hours between two timestamps"""
    try:
        start_dt = pd.to_datetime(start_str, format='%m/%d/%Y %I:%M:%S%p')
        end_dt = pd.to_datetime(end_str, format='%m/%d/%Y %I:%M:%S%p')
        
        if start_dt > end_dt:
            logger.warning(f"Start time {start_str} is after end time {end_str}")
            return 0.0, 0.0
            
        regular_hours = 0.0
        night_hours = 0.0
        current = start_dt
        
        while current < end_dt:
            day_start = datetime.combine(current.date(), time(6, 0))  # 6:00 AM
            day_end = datetime.combine(current.date(), time(18, 0))   # 6:00 PM
            
            if current < day_start:
                # Current time is before 6am (night)
                segment_end = min(day_start, end_dt)
                night_hours += (segment_end - current).total_seconds() / 3600
                current = segment_end
            elif current < day_end:
                # Current time is between 6am and 6pm (regular)
                segment_end = min(day_end, end_dt)
                regular_hours += (segment_end - current).total_seconds() / 3600
                current = segment_end
            else:
                # Current time is after 6pm (night)
                next_day_start = day_start + timedelta(days=1)
                segment_end = min(next_day_start, end_dt)
                night_hours += (segment_end - current).total_seconds() / 3600
                current = segment_end
                
        return round(regular_hours, 5), round(night_hours, 5)
    except Exception as e:
        logger.error(f"Error calculating hours for {start_str} to {end_str}: {e}")
        return 0.0, 0.0

def apply_table_formatting(worksheet, start_row, start_col, end_row, end_col, table_name, header_color):
    """Apply formatting to Excel table"""
    # Header formatting
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting to all columns in the header row
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ensure the header row has consistent formatting
    header_row = worksheet[start_row]
    for cell in header_row:
        if cell.column >= start_col and cell.column <= end_col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

def apply_number_formatting(worksheet, start_row, end_row, start_col, end_col):
    """Apply number formatting to specified columns"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00000'
                cell.alignment = Alignment(horizontal="right")

def add_total_row_formatting(worksheet, row, col_count, fill_color="D3D3D3"):
    """Apply special formatting to total row"""
    total_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    total_font = Font(bold=True)
    total_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, col_count + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = total_alignment
    
    # Ensure the total row has consistent formatting
    total_row = worksheet[row]
    for cell in total_row:
        if cell.column <= col_count:
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = total_alignment

def autofit_columns(worksheet):
    """Auto-fit columns in Excel worksheet"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_time_report(data_df, output_file):
    """Process the time tracking data and generate reports"""
    logger = setup_logging()
    try:
        # Log total agents
        total_agents = data_df['Employee Name'].nunique()
        logger.info(f"Total agents in dataset: {total_agents}")
        
        # Define categories of interest
        categories = [
            'Logged In',
            'Logged Out',
            'Meeting',
            'Meeting - Pre-Shift',
            'Overtime Withdrawals',
            'Withdrawals',
            'Wrap-Up',
            'Training'
        ]
        
        # Filter for selected categories and specific queues only
        filtered_df = data_df[
            data_df['Actual State'].isin(categories) & 
            (data_df['Queue'].isin(['Customer Support', 'BNS', 'IBC Support']))
        ].copy()
        logger.info(f"Filtered to {len(filtered_df)} records in specified categories and queues")
        
        # Extract date from Start time with explicit format
        filtered_df['Date'] = pd.to_datetime(filtered_df['Start'], format='%m/%d/%Y %I:%M:%S%p').dt.date
        
        # Classify IBC vs Non-IBC
        def classify_queue(q):
            if pd.isna(q) or q == '':
                return 'Non IBC'
            return 'IBC' if 'ibc' in str(q).lower() else 'Non IBC'
            
        filtered_df['Queue_Type'] = filtered_df['Queue'].apply(classify_queue)
        
        # Calculate regular and night hours
        logger.info("Calculating regular and night hours")
        regular_hours_list = []
        night_hours_list = []
        
        for index, row in filtered_df.iterrows():
            regular_hrs, night_hrs = calculate_hours(row['Start'], row['Stop'], logger)
            regular_hours_list.append(regular_hrs)
            night_hours_list.append(night_hrs)
            
        filtered_df['Regular Hours'] = regular_hours_list
        filtered_df['Night Hours'] = night_hours_list
        filtered_df['Total Hours'] = filtered_df['Regular Hours'] + filtered_df['Night Hours']
        
        # Create Agent Summary
        logger.info("Creating Agent Summary")
        agent_summary = filtered_df.groupby([
            'Employee Name', 
            'Batch',
            'Queue',
            'Date',
            'Actual State',
            'Queue_Type'
        ]).agg({
            'Regular Hours': 'sum',
            'Night Hours': 'sum',
            'Total Hours': 'sum'
        }).reset_index()
        
        # Separate IBC and Non-IBC summaries
        ibc_agent_summary = agent_summary[agent_summary['Queue_Type'] == 'IBC'].drop(columns='Queue_Type')
        non_ibc_agent_summary = agent_summary[agent_summary['Queue_Type'] == 'Non IBC'].drop(columns='Queue_Type')
        
        # Add totals
        ibc_total = pd.DataFrame({
            'Employee Name': ['TOTAL'],
            'Batch': [''],
            'Queue': [''],
            'Date': [''],
            'Actual State': [''],
            'Regular Hours': [ibc_agent_summary['Regular Hours'].sum()],
            'Night Hours': [ibc_agent_summary['Night Hours'].sum()]
        })
        ibc_agent_summary = pd.concat([ibc_agent_summary, ibc_total], ignore_index=True)
        
        non_ibc_total = pd.DataFrame({
            'Employee Name': ['TOTAL'],
            'Batch': [''],
            'Queue': [''],
            'Date': [''],
            'Actual State': [''],
            'Regular Hours': [non_ibc_agent_summary['Regular Hours'].sum()],
            'Night Hours': [non_ibc_agent_summary['Night Hours'].sum()]
        })
        non_ibc_agent_summary = pd.concat([non_ibc_agent_summary, non_ibc_total], ignore_index=True)
        
        # Create Category Summary
        logger.info("Creating Category Summary")
        category_summary = filtered_df.groupby([
            'Queue_Type',
            'Actual State'
        ]).agg({
            'Regular Hours': 'sum',
            'Night Hours': 'sum'
        }).reset_index()
        
        # Pivot to get IBC and Non-IBC side by side
        ibc_cat = category_summary[category_summary['Queue_Type'] == 'IBC'].drop(columns='Queue_Type')
        non_ibc_cat = category_summary[category_summary['Queue_Type'] == 'Non IBC'].drop(columns='Queue_Type')
        
        final_category = pd.merge(
            ibc_cat,
            non_ibc_cat,
            on='Actual State',
            how='outer',
            suffixes=('_IBC', '_Non_IBC')
        ).fillna(0)
        
        # Add totals row
        cat_total = pd.DataFrame({
            'Actual State': ['TOTAL'],
            'Regular Hours_IBC': [final_category['Regular Hours_IBC'].sum()],
            'Night Hours_IBC': [final_category['Night Hours_IBC'].sum()],
            'Regular Hours_Non_IBC': [final_category['Regular Hours_Non_IBC'].sum()],
            'Night Hours_Non_IBC': [final_category['Night Hours_Non_IBC'].sum()]
        })
        final_category = pd.concat([final_category, cat_total], ignore_index=True)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # IBC Agent Summary Sheet
        ws_ibc = wb.active
        ws_ibc.title = "IBC Agent Summary"
        ws_ibc.cell(row=1, column=1, value="IBC Agent Summary").font = Font(size=14, bold=True)
        
        # Write IBC agent data
        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(ibc_agent_summary, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws_ibc.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        end_row = start_row + len(ibc_agent_summary)
        apply_table_formatting(ws_ibc, start_row, 1, end_row - 1, 7, "IBC_Table", "4472C4")
        apply_number_formatting(ws_ibc, start_row + 1, end_row, 5, 7)  # Format number columns
        add_total_row_formatting(ws_ibc, end_row, 7)
        
        # Non-IBC Agent Summary Sheet
        ws_non_ibc = wb.create_sheet("Non-IBC Agent Summary")
        ws_non_ibc.cell(row=1, column=1, value="Non-IBC Agent Summary").font = Font(size=14, bold=True)
        
        # Write Non-IBC agent data
        start_row_non_ibc = 3
        for r_idx, row in enumerate(dataframe_to_rows(non_ibc_agent_summary, index=False, header=True), start_row_non_ibc):
            for c_idx, value in enumerate(row, 1):
                ws_non_ibc.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        end_row_non_ibc = start_row_non_ibc + len(non_ibc_agent_summary)
        apply_table_formatting(ws_non_ibc, start_row_non_ibc, 1, end_row_non_ibc - 1, 7, "NonIBC_Table", "ED7D31")
        apply_number_formatting(ws_non_ibc, start_row_non_ibc + 1, end_row_non_ibc, 5, 7)  # Format number columns
        add_total_row_formatting(ws_non_ibc, end_row_non_ibc, 7)
        
        # Category Summary Sheet
        ws_cat = wb.create_sheet("Category Summary")
        ws_cat.cell(row=1, column=1, value="Category Summary").font = Font(size=14, bold=True)
        
        # Write category data
        start_row_cat = 3
        for r_idx, row in enumerate(dataframe_to_rows(final_category, index=False, header=True), start_row_cat):
            for c_idx, value in enumerate(row, 1):
                ws_cat.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        cat_end_row = start_row_cat + len(final_category)
        apply_table_formatting(ws_cat, start_row_cat, 1, cat_end_row - 1, 5, "Category_Table", "70AD47")
        apply_number_formatting(ws_cat, start_row_cat + 1, cat_end_row, 2, 5)  # Format number columns
        add_total_row_formatting(ws_cat, cat_end_row, 5, "A9D08E")

         # Create a sheet with all Data for review
        logger.info("Creating All Data sheet for review")

        # Create All Data sheet
        ws_all_data = wb.create_sheet("All Data")
        ws_all_data.cell(row=1, column=1, value="All Data").font = Font(size=14, bold=True)

        # Write all data
        start_row_all_data = 3
        for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=True), start_row_all_data):
            for c_idx, value in enumerate(row, 1):
                ws_all_data.cell(row=r_idx, column=c_idx, value=value)

        # Apply formatting
        all_data_end_row = start_row_all_data + len(data_df)
        apply_table_formatting(ws_all_data, start_row_all_data, 1, all_data_end_row - 1, 8, "All_Data_Table", "4472C4")
        apply_number_formatting(ws_all_data, start_row_all_data + 1, all_data_end_row, 5, 8)  # Format number columns
        add_total_row_formatting(ws_all_data, all_data_end_row, 8)
        
        # Get all unique Actual State categories with their totals
        all_categories_summary = filtered_df.groupby(['Actual State']).agg({
            'Regular Hours': 'sum',
            'Night Hours': 'sum',
            'Total Hours': 'sum'
        }).reset_index()
        # Auto-fit columns
        for sheet in wb.sheetnames:
            autofit_columns(wb[sheet])
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Report saved to {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error processing report: {e}")
        raise

def main():
    parser = argparse.ArgumentParser(description='Process time tracking data and generate reports')
    parser.add_argument('input_file', help='Path to the input time tracking file (CSV or Excel)')
    parser.add_argument('roster_file', help='Path to the employee roster file (Excel)')
    parser.add_argument('-o', '--output', help='Output file path (optional)', default='time_report_complete_output.xlsx')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose logging')
    parser.add_argument('--min-confidence', type=int, default=80, 
                       help='Minimum confidence percentage for name matching (default: 80)')
    
    args = parser.parse_args()
    
    if not os.path.isfile(args.input_file):
        print(f"Error: Input file '{args.input_file}' not found.")
        return
    
    if not os.path.isfile(args.roster_file):
        print(f"Error: Roster file '{args.roster_file}' not found.")
        return
    
    # Set logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    else:
        logging.getLogger().setLevel(logging.INFO)
    
    # Step 1: Process the raw time tracking data
    processed_data = process_employee_data(args.input_file)
    
    if processed_data is None or processed_data.empty:
        print("Error: No valid data found in input file.")
        return
    
    # Step 2: Enrich with employee information from roster (without saving intermediate files)
    try:
        enriched_data, match_rate = add_queue_to_tracking_df(
            processed_data, 
            args.roster_file, 
            args.min_confidence
        )
        print(f"Name matching success rate: {match_rate:.2f}%")
        
        # Step 3: Generate the final report
        generate_time_report(enriched_data, args.output)
        print(f"Successfully created report: {args.output}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        print("Check time_report_complete.log for details")
        exit(1)

if __name__ == "__main__":
    main()
